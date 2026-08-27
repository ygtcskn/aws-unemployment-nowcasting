"""Create the common-sample model comparison workbook and figure."""

from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUTS_DIR = PROJECT_ROOT / "reports" / "outputs"
DEFAULT_TABLES_DIR = PROJECT_ROOT / "reports" / "tables"
DEFAULT_FIGURES_DIR = PROJECT_ROOT / "reports" / "figures"
DEFAULT_MODEL_PANEL = PROJECT_ROOT / "data" / "final" / "gt_unemp_monthly_final.csv"

KEY_COLUMNS = ["date", "country"]
PREDICTION_COLUMNS = ["date", "country", "model", "actual", "prediction"]
LEVEL_LAG_COLUMN = "unemployment_rate_lag1"
TARGET_COLUMN = "unemployment_change_1m"
COUNTRY_PLOTS = {"DE": "Germany", "US": "United States"}

NAVY = "16324F"
BLUE = "2E6F9E"
LIGHT_BLUE = "DCEAF4"
PALE_BLUE = "F2F7FA"
GREEN = "3A7D6B"
AMBER = "C98B2E"
INK = "1F2933"
MUTED = "5F6B76"
LINE = "CBD5DF"
WHITE = "FFFFFF"


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--outputs-dir", type=Path, default=DEFAULT_OUTPUTS_DIR)
    parser.add_argument("--tables-dir", type=Path, default=DEFAULT_TABLES_DIR)
    parser.add_argument("--figures-dir", type=Path, default=DEFAULT_FIGURES_DIR)
    parser.add_argument("--model-panel", type=Path, default=DEFAULT_MODEL_PANEL)
    parser.add_argument(
        "--qa-dir",
        type=Path,
        help="Optional folder for temporary visual-check previews.",
    )
    return parser.parse_args()


# Common-sample preparation

def load_predictions(outputs_dir):
    files = sorted(Path(outputs_dir).rglob("*_predictions.csv"))
    if not files:
        raise FileNotFoundError(
            f"No model prediction files were found under {Path(outputs_dir).resolve()}."
        )

    frames = []
    for path in files:
        frame = pd.read_csv(path)
        missing = set(PREDICTION_COLUMNS) - set(frame.columns)
        if missing:
            raise ValueError(f"{path.name} is missing columns: {sorted(missing)}")

        frame = frame[PREDICTION_COLUMNS].copy()
        frame["date"] = pd.to_datetime(frame["date"], errors="raise")
        frame["country"] = frame["country"].astype(str)
        frame["model"] = frame["model"].astype(str)
        frame["source_file"] = str(path.relative_to(outputs_dir))
        frames.append(frame)

    predictions = pd.concat(frames, ignore_index=True)
    numeric = predictions[["actual", "prediction"]].to_numpy(dtype=float)
    if not np.isfinite(numeric).all():
        raise ValueError("Prediction files contain missing or non-finite values.")

    duplicate = predictions.duplicated(KEY_COLUMNS + ["model"], keep=False)
    if duplicate.any():
        example = predictions.loc[duplicate, KEY_COLUMNS + ["model", "source_file"]]
        raise ValueError(
            "A model has duplicate country-month predictions. First duplicates:\n"
            f"{example.head(10).to_string(index=False)}"
        )
    return predictions, files


def keep_common_sample(predictions):
    models = sorted(predictions["model"].unique())
    if len(models) < 2:
        raise ValueError("At least two models are needed for a comparison.")

    coverage = predictions.groupby(KEY_COLUMNS)["model"].nunique()
    common_index = coverage[coverage.eq(len(models))].index
    keyed = predictions.set_index(KEY_COLUMNS)
    common = keyed.loc[keyed.index.isin(common_index)].reset_index()
    if common.empty:
        raise ValueError("The model outputs do not share any country-month observations.")

    spread = common.groupby(KEY_COLUMNS)["actual"].agg(
        lambda values: values.max() - values.min()
    )
    if spread.gt(1e-9).any():
        bad_key = spread.idxmax()
        raise ValueError(
            "Models disagree on the realized target for "
            f"{bad_key[1]} in {bad_key[0]:%Y-%m}."
        )

    counts = common.groupby("model").size()
    if counts.nunique() != 1:
        raise ValueError("The common-sample filtering produced unequal model samples.")
    return common


def attach_unemployment_levels(predictions, model_panel):
    panel = pd.read_csv(
        model_panel,
        usecols=KEY_COLUMNS + [LEVEL_LAG_COLUMN, TARGET_COLUMN],
        parse_dates=["date"],
    )
    if panel.duplicated(KEY_COLUMNS).any():
        raise ValueError("The modeling panel contains duplicate country-months.")

    merged = predictions.merge(panel, on=KEY_COLUMNS, how="left", validate="many_to_one")
    if merged[[LEVEL_LAG_COLUMN, TARGET_COLUMN]].isna().any().any():
        raise ValueError("Some predictions could not be matched to the modeling panel.")
    if not np.allclose(merged["actual"], merged[TARGET_COLUMN], atol=1e-9):
        raise ValueError("Prediction targets do not match unemployment changes in the panel.")

    merged["actual_level"] = merged[LEVEL_LAG_COLUMN] + merged["actual"]
    merged["prediction_level"] = merged[LEVEL_LAG_COLUMN] + merged["prediction"]
    if merged["actual_level"].le(0).any():
        raise ValueError("MAPE requires positive reconstructed unemployment rates.")
    return merged


def calculate_metrics(predictions):
    rows = []
    for model, group in predictions.groupby("model", sort=True, observed=True):
        error = group["actual"] - group["prediction"]
        rows.append(
            {
                "model": model,
                "rmse": float(np.sqrt(np.mean(np.square(error)))),
                "mae": float(np.mean(np.abs(error))),
                "mape": float(
                    np.mean(
                        np.abs(group["actual_level"] - group["prediction_level"])
                        / group["actual_level"]
                    )
                ),
                "n": int(len(group)),
            }
        )
    return pd.DataFrame(rows).sort_values(["rmse", "model"]).reset_index(drop=True)


# Excel report

def style_title(sheet, cell_range, title):
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[cell.row].height = 32


def style_header(row):
    bottom = Side(style="medium", color=NAVY)
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=bottom)


def write_predictions_sheet(sheet, predictions):
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = BLUE
    style_title(sheet, "A1:K1", "Common-Sample Forecast Data")
    sheet.merge_cells("A2:K2")
    sheet["A2"] = (
        "Models forecast the monthly change in unemployment. Rate levels are "
        "reconstructed from the previous month's rate for a meaningful percentage error."
    )
    sheet["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    sheet["A2"].font = Font(name="Aptos", size=10, italic=True, color=MUTED)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 32

    headers = [
        "Date",
        "Country",
        "Model",
        "Actual Δu",
        "Forecast Δu",
        "Unemployment rate t-1",
        "Actual unemployment rate",
        "Forecast unemployment rate",
        "Absolute error",
        "Squared error",
        "Absolute % error (level)",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    style_header(sheet[4])
    sheet.row_dimensions[4].height = 30

    model_rows = {}
    current_row = 5
    for model, group in predictions.groupby("model", sort=False, observed=True):
        first_row = current_row
        for item in group.itertuples(index=False):
            sheet.cell(current_row, 1, item.date.to_pydatetime())
            sheet.cell(current_row, 2, item.country)
            sheet.cell(current_row, 3, str(item.model))
            sheet.cell(current_row, 4, float(item.actual))
            sheet.cell(current_row, 5, float(item.prediction))
            sheet.cell(current_row, 6, float(item.unemployment_rate_lag1))
            sheet.cell(current_row, 7, f"=D{current_row}+F{current_row}")
            sheet.cell(current_row, 8, f"=E{current_row}+F{current_row}")
            sheet.cell(current_row, 9, f"=ABS(D{current_row}-E{current_row})")
            sheet.cell(current_row, 10, f"=(D{current_row}-E{current_row})^2")
            sheet.cell(
                current_row,
                11,
                f'=IF(ABS(G{current_row})<=1E-8,"",'
                f"ABS((G{current_row}-H{current_row})/G{current_row}))",
            )
            current_row += 1
        model_rows[str(model)] = (first_row, current_row - 1)

    last_row = current_row - 1
    for row in sheet.iter_rows(min_row=5, max_row=last_row, min_col=1, max_col=11):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=INK)
        row[0].number_format = "yyyy-mm-dd"
        for cell in row[3:10]:
            cell.number_format = "0.0000"
        row[10].number_format = "0.00%"

    table = Table(displayName="PredictionsTable", ref=f"A4:K{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "D5"
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 13
    sheet.column_dimensions["E"].width = 13
    sheet.column_dimensions["F"].width = 19
    sheet.column_dimensions["G"].width = 21
    sheet.column_dimensions["H"].width = 22
    sheet.column_dimensions["I"].width = 14
    sheet.column_dimensions["J"].width = 14
    sheet.column_dimensions["K"].width = 20
    return model_rows, last_row


def write_comparison_sheet(sheet, metrics, model_rows):
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = NAVY
    style_title(sheet, "A1:G1", "Monthly Unemployment Nowcasting: Model Comparison")

    sheet.merge_cells("A2:G2")
    sheet["A2"] = (
        "Common out-of-sample country-months. Lower values are better for all "
        "three error measures."
    )
    sheet["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A2"].font = Font(name="Aptos", size=10, color=INK)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 24

    sheet.merge_cells("A4:G4")
    sheet["A4"] = (
        "RMSE and MAE evaluate the forecasted monthly change (percentage points). "
        "MAPE evaluates the reconstructed unemployment-rate level, avoiding "
        "division by zero when Δu = 0."
    )
    sheet["A4"].font = Font(name="Aptos", size=10, italic=True, color=MUTED)
    sheet["A4"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[4].height = 32

    headers = ["Rank", "Model", "RMSE", "MAE", "MAPE", "N", "Evaluation basis"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(6, column, header)
    style_header(sheet[6])
    sheet.row_dimensions[6].height = 30

    summary_start = 7
    summary_end = summary_start + len(metrics) - 1
    thin = Side(style="thin", color=LINE)
    for index, metric in metrics.iterrows():
        row = summary_start + index
        model = metric["model"]
        prediction_start, prediction_end = model_rows[model]
        sheet.cell(row, 1, f"=RANK(C{row},$C${summary_start}:$C${summary_end},1)")
        sheet.cell(row, 2, model)
        sheet.cell(
            row,
            3,
            f"=SQRT(AVERAGE('Predictions'!$J${prediction_start}:$J${prediction_end}))",
        )
        sheet.cell(
            row,
            4,
            f"=AVERAGE('Predictions'!$I${prediction_start}:$I${prediction_end})",
        )
        sheet.cell(
            row,
            5,
            f"=AVERAGE('Predictions'!$K${prediction_start}:$K${prediction_end})",
        )
        sheet.cell(
            row,
            6,
            f"=COUNT('Predictions'!$D${prediction_start}:$D${prediction_end})",
        )
        sheet.cell(row, 7, "Common sample")
        for cell in sheet[row]:
            cell.font = Font(name="Aptos", size=11, color=INK)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)
        sheet.cell(row, 1).alignment = Alignment(horizontal="center")
        for column in range(3, 7):
            sheet.cell(row, column).alignment = Alignment(horizontal="right")
        sheet.cell(row, 1).number_format = "0"
        sheet.cell(row, 3).number_format = "0.0000"
        sheet.cell(row, 4).number_format = "0.0000"
        sheet.cell(row, 5).number_format = "0.00%"
        sheet.cell(row, 6).number_format = "#,##0"
        sheet.row_dimensions[row].height = 23

    for column in ("C", "D", "E"):
        sheet.conditional_formatting.add(
            f"{column}{summary_start}:{column}{summary_end}",
            ColorScaleRule(
                start_type="min",
                start_color="DCEFE8",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF3CD",
                end_type="max",
                end_color="F4CCCC",
            ),
        )

    widths = {"A": 8, "B": 22, "C": 14, "D": 14, "E": 14, "F": 10, "G": 20}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A6:G{summary_end}"
    return summary_start, summary_end


def build_workbook(predictions, metrics, workbook_path):
    workbook = Workbook()
    comparison = workbook.active
    comparison.title = "Model Comparison"
    prediction_sheet = workbook.create_sheet("Predictions")

    model_rows, prediction_end = write_predictions_sheet(
        prediction_sheet, predictions
    )
    summary_start, summary_end = write_comparison_sheet(
        comparison, metrics, model_rows
    )

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(workbook_path)
    return summary_start, summary_end, prediction_end


def validate_workbook(workbook_path, model_count, prediction_count):
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if workbook.sheetnames != ["Model Comparison", "Predictions"]:
        raise AssertionError(f"Unexpected workbook sheets: {workbook.sheetnames}")

    comparison = workbook["Model Comparison"]
    predictions = workbook["Predictions"]
    summary_end = 6 + model_count
    expected_prediction_end = 4 + prediction_count
    if comparison.max_row != summary_end:
        raise AssertionError("The model-comparison table has an unexpected row count.")
    if predictions.max_row != expected_prediction_end:
        raise AssertionError("The prediction audit sheet has an unexpected row count.")

    for row in range(7, summary_end + 1):
        for column in (1, 3, 4, 5, 6):
            value = comparison.cell(row, column).value
            if not isinstance(value, str) or not value.startswith("="):
                raise AssertionError(
                    f"Expected a formula in Model Comparison row {row}, column {column}."
                )

    error_tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(
                    token in cell.value for token in error_tokens
                ):
                    raise AssertionError(f"Formula error token found in {sheet.title}!{cell.coordinate}")
    workbook.close()


# Standalone figure and QA previews

def create_metric_figure(metrics, figure_path):
    plot_data = metrics.copy()
    plot_data["mape"] *= 100
    colors = [f"#{BLUE}", f"#{GREEN}", f"#{AMBER}"]
    specs = [
        ("rmse", "RMSE (percentage points)"),
        ("mae", "MAE (percentage points)"),
        ("mape", "MAPE (unemployment-rate level)"),
    ]

    figure, axes = plt.subplots(1, 3, figsize=(15, 5.2))
    figure.patch.set_facecolor("white")
    for axis, (column, title), color in zip(axes, specs, colors):
        bars = axis.barh(plot_data["model"], plot_data[column], color=color, height=0.58)
        axis.invert_yaxis()
        axis.set_title(title, fontsize=12, weight="bold", color=f"#{INK}")
        axis.grid(axis="x", linestyle="--", alpha=0.35)
        axis.set_axisbelow(True)
        axis.spines[["top", "right", "left"]].set_visible(False)
        axis.tick_params(axis="both", labelsize=9, colors=f"#{MUTED}")
        for bar, value in zip(bars, plot_data[column]):
            display = f"{value:.2f}%" if column == "mape" else f"{value:.4f}"
            axis.text(
                bar.get_width(),
                bar.get_y() + bar.get_height() / 2,
                f"  {display}",
                va="center",
                fontsize=8,
                color=f"#{INK}",
            )

    figure.suptitle(
        "Forecast Accuracy Across Models",
        fontsize=17,
        weight="bold",
        color=f"#{NAVY}",
        y=0.96,
    )
    figure.subplots_adjust(left=0.11, right=0.98, top=0.82, bottom=0.15, wspace=0.42)
    figure.text(
        0.5,
        0.035,
        "Common out-of-sample observations; lower values are better. "
        "MAPE uses reconstructed unemployment-rate levels.",
        ha="center",
        fontsize=9,
        color=f"#{MUTED}",
    )
    figure.savefig(figure_path, dpi=180, bbox_inches="tight")
    plt.close(figure)


def create_country_forecast_figure(predictions, country_code, country_name, path):
    country = predictions[predictions["country"].eq(country_code)].copy()
    if country.empty:
        raise ValueError(f"No common-sample predictions are available for {country_code}.")

    models = [
        str(model)
        for model in predictions["model"].cat.categories
        if country["model"].eq(model).any()
    ]
    actual_spread = country.groupby("date")["actual_level"].agg(
        lambda values: values.max() - values.min()
    )
    if actual_spread.gt(1e-9).any():
        raise ValueError(f"Models disagree on the actual unemployment rate for {country_code}.")

    plotted_values = country[["actual_level", "prediction_level"]].to_numpy()
    lower = float(np.nanmin(plotted_values))
    upper = float(np.nanmax(plotted_values))
    margin = max((upper - lower) * 0.08, 0.25)
    model_colors = [f"#{BLUE}", f"#{GREEN}", f"#{AMBER}", "#8B5E83", "#607D8B"]

    column_count = 2
    row_count = (len(models) + column_count - 1) // column_count
    figure, axes = plt.subplots(
        row_count,
        column_count,
        figsize=(15, 3.4 * row_count + 1.2),
        sharex=True,
        sharey=True,
        squeeze=False,
    )
    axes = axes.ravel()
    for axis, model, color in zip(axes, models, model_colors):
        sample = country[country["model"].eq(model)].sort_values("date")
        axis.plot(
            sample["date"],
            sample["actual_level"],
            color=f"#{INK}",
            linewidth=1.8,
            label="Actual",
            zorder=3,
        )
        axis.plot(
            sample["date"],
            sample["prediction_level"],
            color=color,
            linewidth=1.25,
            alpha=0.9,
            label="Forecast",
            zorder=2,
        )
        axis.set_title(model, fontsize=12, weight="bold", color=f"#{INK}")
        axis.set_ylim(lower - margin, upper + margin)
        axis.grid(True, linestyle="--", alpha=0.3)
        axis.set_axisbelow(True)
        axis.spines[["top", "right"]].set_visible(False)
        axis.tick_params(axis="both", labelsize=9, colors=f"#{MUTED}")
        axis.xaxis.set_major_locator(mdates.YearLocator(3))
        axis.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
        axis.legend(frameon=False, loc="upper left", ncol=2, fontsize=8)

    for axis in axes[len(models):]:
        axis.axis("off")
    for axis in axes[::2]:
        if axis.axison:
            axis.set_ylabel("Unemployment rate (%)", fontsize=9, color=f"#{MUTED}")

    figure.suptitle(
        f"{country_name}: Actual vs Forecast Unemployment Rate",
        fontsize=18,
        weight="bold",
        color=f"#{NAVY}",
        y=0.98,
    )
    figure.text(
        0.5,
        0.025,
        "One-month-ahead rate forecasts reconstructed as "
        "u(t-1) + predicted monthly change. All panels use the common sample.",
        ha="center",
        fontsize=9,
        color=f"#{MUTED}",
    )
    figure.subplots_adjust(left=0.08, right=0.98, top=0.91, bottom=0.08, hspace=0.24)
    figure.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(figure)


def create_qa_previews(metrics, predictions, qa_dir):
    qa_dir.mkdir(parents=True, exist_ok=True)
    display = metrics.copy()
    display.insert(0, "rank", np.arange(1, len(display) + 1))
    display["rmse"] = display["rmse"].map(lambda value: f"{value:.4f}")
    display["mae"] = display["mae"].map(lambda value: f"{value:.4f}")
    display["mape"] = display["mape"].map(lambda value: f"{value:.2%}")
    display["n"] = display["n"].map(lambda value: f"{value:,}")

    figure, axis = plt.subplots(figsize=(11, 3.4))
    axis.axis("off")
    axis.set_title(
        "Monthly Unemployment Nowcasting: Model Comparison",
        loc="left",
        fontsize=16,
        weight="bold",
        color=f"#{NAVY}",
        pad=18,
    )
    table = axis.table(
        cellText=display.values,
        colLabels=["Rank", "Model", "RMSE", "MAE", "MAPE", "N"],
        cellLoc="center",
        colLoc="center",
        loc="upper left",
        bbox=[0, 0.05, 1, 0.75],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    for (row, _), cell in table.get_celld().items():
        if row == 0:
            cell.set_facecolor(f"#{BLUE}")
            cell.set_text_props(color="white", weight="bold")
        else:
            cell.set_edgecolor(f"#{LINE}")
    figure.savefig(qa_dir / "model_comparison_preview.png", dpi=160, bbox_inches="tight")
    plt.close(figure)

    sample = predictions.head(10).copy()
    sample["date"] = sample["date"].dt.strftime("%Y-%m-%d")
    sample = sample[["date", "country", "model", "actual", "prediction", "actual_level"]]
    sample[["actual", "prediction", "actual_level"]] = sample[
        ["actual", "prediction", "actual_level"]
    ].map(lambda value: f"{value:.4f}")
    figure, axis = plt.subplots(figsize=(13, 4.2))
    axis.axis("off")
    axis.set_title(
        "Prediction Audit Sheet — First 10 Rows",
        loc="left",
        fontsize=15,
        weight="bold",
        color=f"#{NAVY}",
        pad=15,
    )
    table = axis.table(
        cellText=sample.values,
        colLabels=["Date", "Country", "Model", "Actual Δu", "Forecast Δu", "Actual rate"],
        cellLoc="center",
        colLoc="center",
        loc="upper left",
        bbox=[0, 0.02, 1, 0.82],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    for (row, _), cell in table.get_celld().items():
        if row == 0:
            cell.set_facecolor(f"#{BLUE}")
            cell.set_text_props(color="white", weight="bold")
        else:
            cell.set_edgecolor(f"#{LINE}")
    figure.savefig(qa_dir / "predictions_preview.png", dpi=160, bbox_inches="tight")
    plt.close(figure)


def main():
    args = parse_args()
    outputs_dir = args.outputs_dir.resolve()
    tables_dir = args.tables_dir.resolve()
    figures_dir = args.figures_dir.resolve()
    model_panel = args.model_panel.resolve()

    predictions, files = load_predictions(outputs_dir)
    predictions = keep_common_sample(predictions)
    predictions = attach_unemployment_levels(predictions, model_panel)
    metrics = calculate_metrics(predictions)

    model_order = metrics["model"].tolist()
    predictions["model"] = pd.Categorical(
        predictions["model"], categories=model_order, ordered=True
    )
    predictions = predictions.sort_values(["model", "date", "country"])

    tables_dir.mkdir(parents=True, exist_ok=True)
    figures_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = tables_dir / "model_comparison.xlsx"
    figure_path = figures_dir / "model_metric_comparison.png"

    build_workbook(predictions, metrics, workbook_path)
    validate_workbook(workbook_path, len(metrics), len(predictions))
    create_metric_figure(metrics, figure_path)
    country_figure_paths = []
    for country_code, country_name in COUNTRY_PLOTS.items():
        country_path = figures_dir / f"{country_code.lower()}_actual_vs_predicted.png"
        create_country_forecast_figure(
            predictions, country_code, country_name, country_path
        )
        country_figure_paths.append(country_path)
    if args.qa_dir:
        create_qa_previews(metrics, predictions, args.qa_dir.resolve())

    print(f"Loaded {len(files)} prediction files from {outputs_dir}")
    print(f"Common sample: {metrics['n'].iloc[0]:,} country-months per model")
    print("\nModel comparison (MAPE is evaluated on unemployment-rate levels):\n")
    display = metrics.copy()
    display["mape"] *= 100
    print(display.to_string(index=False, float_format=lambda value: f"{value:.4f}"))
    print(f"\nWorkbook: {workbook_path}")
    print(f"Figure:   {figure_path}")
    for path in country_figure_paths:
        print(f"Figure:   {path}")


if __name__ == "__main__":
    main()
